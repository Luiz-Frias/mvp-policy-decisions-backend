"""Admin quote management endpoints."""

import csv
import io
import json
from datetime import datetime
from typing import Any
from uuid import UUID

from beartype import beartype
from fastapi import APIRouter, Depends, Query, Response
from pydantic import Field

from ....models.admin import AdminUser
from ....models.base import BaseModelConfig
from ....models.quote import QuoteOverrideRequest
from ....schemas.quote import QuoteResponse
from ....services.quote_service import QuoteService
from ...dependencies import get_current_admin_user, get_quote_service
from ...response_patterns import ErrorResponse

# Auto-generated models


@beartype
class ResultsData(BaseModelConfig):
    """Structured model replacing dict[str, Any] usage."""

    # Auto-generated - customize based on usage
    content: str | None = Field(default=None, description="Content data")
    metadata: dict[str, str] = Field(default_factory=dict, description="Metadata")


@beartype
class ParametersData(BaseModelConfig):
    """Structured model replacing dict[str, Any] usage."""

    # Auto-generated - customize based on usage
    content: str | None = Field(default=None, description="Content data")
    metadata: dict[str, str] = Field(default_factory=dict, description="Metadata")


router = APIRouter()


@router.get("/search", response_model=list[QuoteResponse])
@beartype
async def admin_search_quotes(
    response: Response,
    # Search filters
    status: str | None = None,
    state: str | None = None,
    min_premium: float | None = None,
    max_premium: float | None = None,
    created_after: datetime | None = None,
    created_before: datetime | None = None,
    customer_email: str | None = None,
    # Pagination
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    # PII control
    include_pii: bool = Query(False),
    # Dependencies
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> list[dict[str, Any]] | ErrorResponse:
    """Search quotes with admin privileges."""
    filters = {
        "status": status,
        "state": state,
        "min_premium": min_premium,
        "max_premium": max_premium,
        "created_after": created_after,
        "created_before": created_before,
        "customer_email": customer_email,
    }

    result = await quote_service.admin_search_quotes(
        admin_user.id,
        filters,
        include_pii,
    )

    if result.is_err():
        error_msg = result.unwrap_err()
        response.status_code = 400 if "invalid" in error_msg.lower() else 500
        return ErrorResponse(error=error_msg)

    # Ensure we return a valid list, not None
    value = result.unwrap()
    # Convert Quote objects to dict format for JSON response
    return [quote.model_dump() if hasattr(quote, "model_dump") else quote for quote in value]  # type: ignore[misc]


@router.post("/{quote_id}/override")
@beartype
async def override_quote(
    quote_id: UUID,
    override_request: QuoteOverrideRequest,
    response: Response,
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> dict[str, Any] | ErrorResponse:
    """Override quote pricing or terms."""
    # Check permission
    if "quote:override" not in admin_user.effective_permissions:
        response.status_code = 403
        return ErrorResponse(error="Insufficient permissions")

    result = await quote_service.admin_override_quote(
        quote_id,
        admin_user.id,
        override_request,
    )

    if result.is_err():
        error_msg = result.unwrap_err()
        response.status_code = 404 if "not found" in error_msg.lower() else 400
        return ErrorResponse(error=error_msg)

    # Ensure we return a valid quote, not None
    value = result.unwrap()
    # Convert Quote object to dict format for JSON response
    return value.model_dump() if hasattr(value, "model_dump") else dict(value)


@router.post("/bulk/{operation}")
@beartype
async def bulk_quote_operation(
    operation: str,
    quote_ids: list[UUID],
    response: Response,
    parameters: ParametersData | None = None,
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> dict[str, Any] | ErrorResponse:
    """Perform bulk operations on quotes."""
    # Validate operation
    allowed_operations = ["expire", "extend", "recalculate", "export"]
    if operation not in allowed_operations:
        response.status_code = 400
        return ErrorResponse(
            error=f"Invalid operation. Allowed: {', '.join(allowed_operations)}"
        )

    # Check permissions
    required_permission = f"quote:bulk_{operation}"
    if required_permission not in admin_user.effective_permissions:
        response.status_code = 403
        return ErrorResponse(error="Insufficient permissions")

    # Process in batches
    batch_size = 50
    results: ResultsData = {
        "total": len(quote_ids),
        "successful": 0,
        "failed": 0,
        "errors": [],
    }

    for i in range(0, len(quote_ids), batch_size):
        batch = quote_ids[i : i + batch_size]

        if operation == "expire":
            # Expire quotes
            for quote_id in batch:
                try:
                    await quote_service._db.execute(
                        """
                        UPDATE quotes
                        SET status = 'EXPIRED',
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = $1 AND status != 'BOUND'
                        """,
                        quote_id,
                    )
                    results["successful"] += 1
                except Exception as e:
                    results["failed"] += 1
                    results["errors"].append(
                        {"quote_id": str(quote_id), "error": str(e)}
                    )

        elif operation == "extend":
            # Extend expiration
            days = parameters.get("days", 30) if parameters else 30
            for quote_id in batch:
                try:
                    await quote_service._db.execute(
                        """
                        UPDATE quotes
                        SET expires_at = expires_at + INTERVAL '%s days',
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = $1 AND status NOT IN ('BOUND', 'EXPIRED')
                        """,
                        days,
                        quote_id,
                    )
                    results["successful"] += 1
                except Exception as e:
                    results["failed"] += 1
                    results["errors"].append(
                        {"quote_id": str(quote_id), "error": str(e)}
                    )

        elif operation == "recalculate":
            # Trigger recalculation
            for quote_id in batch:
                result = await quote_service.calculate_quote(quote_id)
                if result.is_ok():
                    results["successful"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append(
                        {"quote_id": str(quote_id), "error": result.err_value}
                    )

    return results


@router.get("/analytics")
@beartype
async def get_quote_analytics(
    date_from: datetime,
    date_to: datetime,
    response: Response,
    group_by: str = Query("day", regex="^(hour|day|week|month)$"),
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> dict[str, Any] | ErrorResponse:
    """Get quote analytics for dashboards."""
    result = await quote_service.get_quote_analytics(
        date_from,
        date_to,
        group_by,
    )

    if result.is_err():
        error_msg = result.unwrap_err()
        response.status_code = 500
        return ErrorResponse(error=error_msg)

    return result.unwrap()


@router.get("/export")
@beartype
async def export_quotes(
    response: Response,
    format: str = Query("csv", regex="^(csv|json|excel)$"),
    status: str | None = None,
    state: str | None = None,
    created_after: datetime | None = None,
    created_before: datetime | None = None,
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> dict[str, Any] | ErrorResponse:
    """Export quotes in various formats."""
    # Check permission
    if "quote:export" not in admin_user.effective_permissions:
        response.status_code = 403
        return ErrorResponse(error="Insufficient permissions")

    # Build query
    query_parts = ["SELECT * FROM quotes WHERE 1=1"]
    params = []
    param_count = 0

    if status:
        param_count += 1
        query_parts.append(f"AND status = ${param_count}")
        params.append(status)

    if state:
        param_count += 1
        query_parts.append(f"AND state = ${param_count}")
        params.append(state)

    if created_after:
        param_count += 1
        query_parts.append(f"AND created_at >= ${param_count}")
        params.append(str(created_after))

    if created_before:
        param_count += 1
        query_parts.append(f"AND created_at <= ${param_count}")
        params.append(str(created_before))

    query = " ".join(query_parts)
    rows = await quote_service._db.fetch(query, *params)

    # Format data based on requested format
    if format == "json":
        data = [dict(row) for row in rows]
        return {"format": "json", "count": len(data), "data": data}
    elif format == "csv":
        # Implement CSV export
        if not rows:
            return {
                "format": "csv",
                "count": 0,
                "data": "",
                "message": "No quotes found for export",
            }

        # Create CSV output
        output = io.StringIO()

        # Get column names from first row
        headers = list(rows[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)

        # Write headers
        writer.writeheader()

        # Write data rows
        for row in rows:
            # Convert row to dict and handle JSON/UUID fields
            row_dict = {}
            for key, value in row.items():
                if isinstance(value, (dict, list)):
                    # Convert complex types to JSON strings
                    row_dict[key] = json.dumps(value)
                elif value is None:
                    row_dict[key] = ""
                else:
                    row_dict[key] = str(value)
            writer.writerow(row_dict)

        csv_content = output.getvalue()
        output.close()

        return {
            "format": "csv",
            "count": len(rows),
            "data": csv_content,
            "filename": f"quotes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        }
    else:
        # Implement Excel export using openpyxl
        try:
            import openpyxl  # type: ignore[import-untyped]
            import pandas as pd  # type: ignore[import-untyped]
            from openpyxl.styles import (  # type: ignore[import-untyped]
                Font,
                PatternFill,
            )
            from openpyxl.utils.dataframe import (
                dataframe_to_rows,  # type: ignore[import-untyped]
            )

        except ImportError:
            # Graceful fallback if Excel libraries not available
            return {
                "format": "excel",
                "count": len(rows),
                "message": "Excel export requires openpyxl and pandas libraries. Please install with: pip install openpyxl pandas",
                "fallback_csv": True,
            }

        if not rows:
            return {
                "format": "excel",
                "count": 0,
                "data": "",
                "message": "No quotes found for export",
            }

        # Convert to DataFrame for easier Excel manipulation
        data = []
        for row in rows:
            row_dict = {}
            for key, value in row.items():
                if isinstance(value, (dict, list)):
                    # Convert complex types to JSON strings for Excel
                    row_dict[key] = json.dumps(value)
                elif value is None:
                    row_dict[key] = ""
                else:
                    row_dict[key] = value
            data.append(row_dict)

        df = pd.DataFrame(data)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quotes Export"

        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        # Write DataFrame to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save to binary buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Convert to base64 for JSON response
        import base64

        excel_b64 = base64.b64encode(excel_buffer.getvalue()).decode("utf-8")

        return {
            "format": "excel",
            "count": len(rows),
            "data": excel_b64,
            "filename": f"quotes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }


@router.get("/approvals/pending")
@beartype
async def get_pending_approvals(
    response: Response,
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> list[dict[str, Any]] | ErrorResponse:
    """Get quotes pending admin approval."""
    # Check permission
    if "quote:approve" not in admin_user.effective_permissions:
        response.status_code = 403
        return ErrorResponse(error="Insufficient permissions")

    query = """
        SELECT
            qa.*,
            q.quote_number,
            q.total_premium,
            q.state,
            q.product_type
        FROM quote_approvals qa
        JOIN quotes q ON qa.quote_id = q.id
        WHERE qa.status = 'pending'
        ORDER BY qa.requested_at DESC
    """

    rows = await quote_service._db.fetch(query)

    return [dict(row) for row in rows]


@router.post("/approvals/{approval_id}/{action}")
@beartype
async def process_approval(
    approval_id: UUID,
    action: str,
    response: Response,
    notes: str | None = None,
    quote_service: QuoteService = Depends(get_quote_service),
    admin_user: AdminUser = Depends(get_current_admin_user),
) -> dict[str, Any] | ErrorResponse:
    """Approve or reject a quote approval request."""
    if action not in ["approve", "reject"]:
        response.status_code = 400
        return ErrorResponse(error="Action must be 'approve' or 'reject'")

    # Check permission
    if "quote:approve" not in admin_user.effective_permissions:
        response.status_code = 403
        return ErrorResponse(error="Insufficient permissions")

    status = "approved" if action == "approve" else "rejected"

    # Update approval
    await quote_service._db.execute(
        """
        UPDATE quote_approvals
        SET status = $2,
            reviewed_by = $3,
            reviewed_at = $4,
            notes = $5
        WHERE id = $1
        """,
        approval_id,
        status,
        admin_user.id,
        datetime.now(),
        notes,
    )

    return {
        "approval_id": str(approval_id),
        "action": action,
        "status": status,
        "reviewed_by": str(admin_user.id),
        "reviewed_at": datetime.now().isoformat(),
    }
